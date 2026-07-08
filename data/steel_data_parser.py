from openpyxl import load_workbook
import json
import re


def parse_section(raw):
    """Parse '90 x 90 x 10' angle section string into components."""
    if not raw:
        return {"raw": None, "leg1_mm": None, "leg2_mm": None, "thickness_mm": None}
    parts = re.split(r"\s*[xX]\s*", str(raw).strip())
    try:
        return {
            "raw": str(raw).strip(),
            "leg1_mm":      float(parts[0]) if len(parts) > 0 else None,
            "leg2_mm":      float(parts[1]) if len(parts) > 1 else None,
            "thickness_mm": float(parts[2]) if len(parts) > 2 else None,
        }
    except ValueError:
        return {"raw": str(raw).strip(), "leg1_mm": None, "leg2_mm": None, "thickness_mm": None}


def parse_dimensions(raw):
    """Parse '3.8 X 3.8' pile cap dimension string."""
    if not raw:
        return {"raw": None, "width_m": None, "length_m": None}
    parts = re.split(r"\s*[xX]\s*", str(raw).strip())
    try:
        return {
            "raw":      str(raw).strip(),
            "width_m":  float(parts[0]) if len(parts) > 0 else None,
            "length_m": float(parts[1]) if len(parts) > 1 else None,
        }
    except ValueError:
        return {"raw": str(raw).strip(), "width_m": None, "length_m": None}


def parse_steel_superstructure_ifc(filepath):
    wb = load_workbook(filepath, read_only=True)
    ws = wb["Steel_Superstructure_IFC"]

    records = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        if row[0] is None:
            continue  # skip empty rows

        # Stiffener ranges
        stiffener_ranges = [
            {"section": "support",      "initial_m": row[47], "end_m": row[49]},
            {"section": "intermediate", "initial_m": row[50], "end_m": row[52]},
            {"section": "mid",          "initial_m": row[53], "end_m": row[55]},
            {"section": "intermediate", "initial_m": row[56], "end_m": row[58]},
            {"section": "support",      "initial_m": row[59], "end_m": row[61]},
        ]
        # Filter out empty/duplicate ranges (e.g. intermediate with 0 length)
        stiffener_ranges = [
            s for s in stiffener_ranges
            if s["initial_m"] is not None and s["end_m"] is not None
        ]

        record = {
            "bridge_id":     row[0],
            "name_display":  row[1],
            "location":      row[2],
            "bridge_type":   row[3],
            "span_m":        row[4],
            "lanes":         row[5],
            "has_footpath":  row[6] == "Yes" if row[6] is not None else False,
            "total_width_m": row[8],
            "material_grades": {
                "rebar":                               row[9],
                "structural_steel":                    row[10],
                "concrete_substructure_and_deck_slab": row[11],
            },
            "deck_slab": {
                "depth_m": row[13],
            },
            "girder_spacing_m": row[15],
            "cross_bracing": {
                "spacing_m": row[16],
                "count":     row[17],
                "section":   parse_section(row[18]),
            },
            "transverse_stiffeners_per_girder_side": {
                "width_m":     row[19],
                "thickness_m": row[20],
                "spacing_m":   row[21],
            },
            "bearing_stiffeners_per_girder_side": {
                "width_m":     row[22],
                "thickness_m": row[23],
                "spacing_m":   row[24],
            },
            "girder_sections": {
                "mid": {
                    "top_flange_width_m":        row[25],
                    "top_flange_thickness_m":    row[26],
                    "web_depth_m":               row[27],
                    "web_thickness_m":           row[28],
                    "bottom_flange_width_m":     row[29],
                    "bottom_flange_thickness_m": row[30],
                    "length_per_girder_m":       row[31],
                },
                "intermediate": {
                    "top_flange_width_m":        row[32],
                    "top_flange_thickness_m":    row[33],
                    "web_depth_m":               row[34],
                    "web_thickness_m":           row[35],
                    "bottom_flange_width_m":     row[36],
                    "bottom_flange_thickness_m": row[37],
                    "length_per_girder_m":       row[38],
                },
                "support": {
                    "top_flange_width_m":        row[39],
                    "top_flange_thickness_m":    row[40],
                    "web_depth_m":               row[41],
                    "web_thickness_m":           row[42],
                    "bottom_flange_width_m":     row[43],
                    "bottom_flange_thickness_m": row[44],
                    "length_per_girder_m":       row[45],
                },
                "end_diaphragm": {
                    "top_flange_width_m":        row[63],
                    "top_flange_thickness_m":    row[64],
                    "web_depth_m":               row[65],
                    "web_thickness_m":           row[66],
                    "bottom_flange_width_m":     row[67],
                    "bottom_flange_thickness_m": row[68],
                },
                "longitudinal_segmentation_m": stiffener_ranges,
            },
            "substructure": {
                "pile": {
                    "count":           row[70],
                    "depth_m":         row[71],
                    "diameter_m":      row[72],
                    "reinforcement_pct": row[73],
                },
                "pile_cap": {
                    "depth_m":          row[74],
                    "dimensions_m":     parse_dimensions(row[75]),
                    "pcc_depth_m":      row[76],
                    "reinforcement_pct": row[77],
                },
                "pier": {
                    "height_m":   row[79],
                    "diameter_m": row[80],
                },
                "pier_cap": {
                    "mid_depth_m":           row[81],
                    "end_depth_m":           row[82],
                    "width_m":               row[83],
                    "mid_section_length_m":  row[84],
                    "taper_section_length_m": row[85],
                    "reinforcement_pct":     row[86],
                },
            }
        }
        records.append(record)

    return records


if __name__ == "__main__":
    import os
    here = os.path.dirname(__file__)
    filepath = os.path.join(here, "Steel_girder_Bridge_Drawings.xlsm")
    records = parse_steel_superstructure_ifc(filepath)
    out_path = os.path.join(here, "steel_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, default=str)
    print(f"Wrote {len(records)} records to {out_path}")